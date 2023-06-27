#!/usr/local/bin/python
# -*- coding: utf-8 -*-
import openpyxl, traceback
import os, os.path, time, json, pathlib, getpass
import requests, lxml, re, math
from bs4 import BeautifulSoup
from random import randrange
from lxml.html import fromstring
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import TimeoutException
from datetime import datetime, date
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import *

########################################################### [ CHORMEDRIVER ] ###########################################################
options = webdriver.ChromeOptions()
options.add_argument("start-maximized")
options.add_argument('disable-extensions')
options.add_argument('disable-gpu')
options.add_argument('disable-infobars')
options.add_argument('disable-ntp-most-likely-favicons-from-server')
options.add_argument('disable-login-animations')
options.add_argument('disable-popup-blocking')
options.add_argument('disable-images')
options.add_argument('log-level=0')
options.add_argument('log-level=1')
options.add_argument('log-level=2')
options.add_argument('log-level=3')
options.add_experimental_option("prefs",{'profile.managed_default_content_settings.images': 2})
# logDriver = webdriver.Chrome(options=options)
# time.sleep(2)
########################################################### [ CHORMEDRIVER ] ###########################################################

########################################################### [ General Content ] ###########################################################
current_directory = os.getcwd()+"\\"
today_date = date.today()
print('Date: ',today_date)

def file_name_checker(name):
    for char in ['@','$','%','&','\\','/',':','*','?','"',"'",'<','>','|','~','`','#','^','+','=','{','}','[',']',';','!']:
        if char in name:
            name = name.replace(char, "_")
    return name

headddy = {
	'cookie': 'bc_cp_user=eyJhbGciOiJkaXIiLCJlbmMiOiJBMTI4Q0JDLUhTMjU2In0..jpLbMbXjiSXrYA55RrAhUg.qkMvHzVwlp149p753XihdlW3XL_rXa50q-07CyWU3Vz2WHWxjfG-HzTWMu3rN_C1AF_HhZL_r2OZl_zGgHB4Vkfz5JM9xl_sTFwiUgvk43bUSSzvpoQetycjUrFh-eDJa1iR7Vj1ZzIgPPjmkpFPQz1fTCbCSm8B9KSB9dfiL5X2yIhJJM0jqA6Bq17l2HC_gucCiGY19eQd4M3dT0HY3yKI_p7cri_8rjSiAkT3BY6Vdu2aTjVUxN8R3bxIwuUCFK7iYBAj8-U_ei1AZJzGZwKmt8C1BajmBRtVyrvTOj9-cCVZ30cepvpDDEXeicfheAMfldIJ_PD6NumCZWQbGwuJBMafTf8pNUCq7_iVioAcCmLjZu-7fBri5N9TJunAWYr8gzt5PagJ6OOBYOcF7ukVJ7nNc7JxuYGwRCUKdcES4Nm6hah3oM0LvjDJUaTh-EaY6_gnLbJjlYtzCze0Q5QYbSB7l7R1J6_Y37uFR7yx2InQxclHbb6qPGFdaZyvwHtuO-k8xZqzGMHegg3YPDS0-RXNSXLJQqNFB5nudmspN01PogeDRVPhTMn-GSIdJFXJvENRM_L3ll2ZQqpjp4yJrlzUEC7151CY8FS2SC-e4Zf1CmAcJIDjf2oxpWlo6Lf_SBgTUewxlUvKkNJizSOFsCdgCbz26Tjv9HIAM6UPqhuU3bfKiJ8MNfXmW3uMZ266MLdBtqD1SarY0jOyDDCXB6G_VFWiO4qsW_2zVKjOyZbf9EmtOnW6b702oUg7T-hRErcyTJvvXIJXzBpgAVNts6vBZ0r3PVc4pHzeEz6s0oMXAWK6GSVRsUbXOWljdw1i6RAtP4CmZ5SUgPXYkW4vlHP8jrq66VgOZmHsQ7outrNisF4s3Eug1BO_ad1WYS1JHUzBHU8_8bXivV2a2JV0b1M3V8wToL9Hl78rE05WU4vVMjtOUSCQnpGVgk6JoAXdCOERNFCDO_wQvIgWdnEIkwFgGL72pGS3xyKvxQ-iRNDMtda1901TQEYO-2IEsRdpE1iCIakL6j55D13_CImMuV3-AF7mzzg6EAxDPNU.aI-aACty7XOOSBxAWRXA1A; ajs_anonymous_id=d1d30f98-2a71-41c4-ba6a-81d9cd735149; ajs_group_id=10091980; ajs_user_id=10091980-43; _gcl_au=1.1.1810766942.1686134319; bcactive=yes; bc_segment_id=10091980-43; __utma=106057409.616101767.1686134319.1686134335.1686134335.1; __utmz=106057409.1686134335.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); __utma=106057409.616101767.1686134319.1686134335.1686134335.1; __utmz=106057409.1686134335.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); bc_storeId=10091980; bc_storeCountry=United States; bc_plan=Enterprise Store Base Monthly; _gid=GA1.2.61371875.1686229249; SHOP_SESSION_TOKEN=d1a33n04pblt16vbdin2n18ppd; bc_auth_session=bc%3Aoauth_session%3A991860f6404f97aa9408ead9b526c9c5e1e23e4ca1d0c47353af4a18d668d724; STORESUITE_CP_TOKEN=z44IFlvNgCIR9jFDM8Fjv2PzQncbZqtSM757InE7; XSRF-TOKEN=e7c38c3271cf641fe350cf4fd3acf13b6262f26ec9fb55f32d595eff37bbac5a; _gat_UA-10308612-26=1; _gat_UA-10308612-33=1; _ga_WS2VZYPC6G=GS1.1.1686295936.5.1.1686295968.28.0.0; _ga=GA1.1.616101767.1686134319',
	'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.50',
	'x-xsrf-token': 'e7c38c3271cf641fe350cf4fd3acf13b6262f26ec9fb55f32d595eff37bbac5a'
}

col_headers_list = ['Category','Name', 'Count/Details', 'Visible']
########################################################### [ General Content ] ###########################################################

#################### workbook module ############################
aits_book = openpyxl.Workbook()
fields_sheet = aits_book.active
fields_sheet.title = 'Categories'

for col_row, col_headers in enumerate(col_headers_list):
	fields_sheet.cell(1, col_row+1).value = col_headers

fields_sheet_row = fields_sheet.max_row+1
#################### workbook module ############################

filter_url = 'https://store-mzpq4daxph.mybigcommerce.com/admin/services/public/stores/mzpq4daxph/v3/settings/search/filters'
avail_url = 'https://store-mzpq4daxph.mybigcommerce.com/admin/services/public/stores/mzpq4daxph/v3/settings/search/filters/available?channel_id=1&category_id='

pager = requests.get(filter_url, headers=headddy)
filter_json = json.loads(pager.text)

# # ################################################ Category - Extraction ################################################
# category_url = 'https://store-mzpq4daxph.mybigcommerce.com/admin/services/public/stores/mzpq4daxph/v3/catalog/trees/1/categories'
# cat_ping = requests.get(category_url, headers=headddy)
# cat_json = json.loads(cat_ping.text)

# category_list = []
# for buckets in cat_json['data']:
# 	for inner_bucket1 in buckets['children']:
# 		second_level = inner_bucket1['name']
# 		second_id = inner_bucket1['id']
# 		category_list.append(str(second_id)+"|"+second_level)
# 		if inner_bucket1['children'] != []:
# 			for inner_bucket2 in inner_bucket1['children']:
# 				third_level = inner_bucket2['name']
# 				third_id = inner_bucket2['id']
# 				category_list.append(str(third_id)+"|"+third_level)
# 				if inner_bucket2['children'] != []:
# 					for inner_bucket3 in inner_bucket2['children']:
# 						fourth_level = inner_bucket3['name']
# 						fourth_id = inner_bucket3['id']
# 						category_list.append(str(fourth_id)+"|"+fourth_level)
# 						if inner_bucket3['children'] != []:
# 							for inner_bucket4 in inner_bucket3['children']:
# 								fifth_level = inner_bucket4['name']
# 								fifth_id = inner_bucket4['id']
# 								category_list.append(str(fifth_id)+"|"+fifth_level)
# 								if inner_bucket4['children'] != []:
# 									for inner_bucket5 in inner_bucket4['children']:
# 										sixth_level = inner_bucket5['name']
# 										sixth_id = inner_bucket5['id']
# 										category_list.append(str(sixth_id)+"|"+sixth_level)
# 										if inner_bucket5['children'] != []:
# 											for inner_bucket6 in inner_bucket5['children']:
# 												seventh_level = inner_bucket6['name']
# 												seventh_id = inner_bucket6['id']
# 												category_list.append(str(seventh_id)+"|"+seventh_level)
# 												if inner_bucket6['children'] != []:
# 													for inner_bucket7 in inner_bucket6['children']:
# 														eighth_level = inner_bucket7['name']
# 														eighth_id = inner_bucket7['id']
# 														category_list.append(str(eighth_id)+"|"+eighth_level)

# for caters in category_list:
# 	fields_sheet.cell(fields_sheet_row, 1).value = caters.split('|')[0]
# 	fields_sheet.cell(fields_sheet_row, 2).value = caters.split('|')[1]
# 	fields_sheet_row+=1
# aits_book.save('temp_categories.xlsx')

################################################ Get Categroy input ################################################
##################### input
get_file = input('Enter the category file name without extensions: ')
path = current_directory+'/'+get_file+'.xlsx'
con_in_wb = openpyxl.load_workbook(path)
con_in_ws = con_in_wb.active
con_max_rows = con_in_ws.max_row+1

input_category_list = [con_in_ws.cell(i, 1).value for i in range(2, con_max_rows)]

################################################ Category - Filters ################################################

# ################################################ Category - Filters ################################################
for cat_num, category_data in enumerate(input_category_list):
	category_id = category_data.split('|')[0]
	category_name = category_data.split('|')[1]
	print(f'Processsing [{cat_num+1}/{len(input_category_list)}] --------------- {category_name} ---------------')
	available_details = []
	a_pager = requests.get(avail_url+category_id, headers=headddy)
	avail_json = json.loads(a_pager.text)
	######################### for Product Count #########################
	for dta in avail_json['data']:
		product_details = ''
		if 'product_count' not in dta:
			if 'price_range_min' in dta:
				product_details = str(dta['price_range_min'])+" - "+str(dta['price_range_max'])
		else:
			product_details = str(dta['product_count'])
		available_details.append((dta['id'], dta['name'], product_details))

	write_list = []
	for categories in filter_json['data']:
		field_id = categories['id']
		field_name = categories['display_name']
		field_visible = "Yes" if categories['is_enabled'] == True else "No"
		[write_list.append((category_name, field_name, border[2], field_visible)) for border in available_details if field_id == border[0]]

	for pretty, wrt_val in enumerate(write_list):
		for ind_wrt, wrt_col in enumerate(wrt_val):
			fields_sheet.cell(fields_sheet_row, ind_wrt+1).value = wrt_col
		fields_sheet_row+=1

	aits_book.save(f'AITS - Category Filters_{today_date}.xlsx')

try:
	aits_book.save(f'AITS - Category Filters_{today_date}.xlsx')
except:
	print('exception..')
	aits_book.save(f'AITS - Category Filters_{today_date}.xlsx')
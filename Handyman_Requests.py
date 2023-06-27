import os,time,re,traceback
import selenium,requests
from datetime import datetime
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver

# driver = webdriver.Chrome(options=options)
#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
path = os.getcwd()
wb = openpyxl.load_workbook(filename =input('Excel File : ')+'.xlsx')

sheet1 = wb.active
sheet1['A1'] = 'Product Url'
sheet1['B1'] = 'Title'
sheet1['C1'] = 'Price (₱)'
sheet1['D1'] = 'Striked Price (₱)'
sheet1['E1'] = 'Brand'
sheet1['F1'] = 'Sku'
sheet1['G1'] = 'Image'
sheet1['H1'] = 'Description'
start_time = datetime.now()
maxi_row = sheet1.max_row+1
data_row = 2
# data_row = int(input('Starting Range : '))
# end_range = int(input('Ending Range : '))
for row_num in range(2,maxi_row):
    '<<<<<<<<<<<<<<<<<<<<-------------Url in Sitemap-------------->>>>>>>>>>>>>>>>>>>>>'
    input_url = sheet1.cell(row_num,1).value
    if input_url != None or input_url != 'None' or input_url != 'none':
        try:
            response = requests.get(input_url)
        except:
            print('Sleeping..........')
            time.sleep(300)
            response = requests.get(input_url)
        soup = BeautifulSoup(response.content, 'html.parser')
        prod_title = soup.select('.product-main .product-title')[0].text.strip() if soup.select('.product-main .product-title') != [] else ''
        prod_price = soup.select('.product-main .price--main span.money')[0].text.strip().replace('₱', '') if soup.select('.product-main .price--main span.money') != [] else ''
        prod_desc = soup.select('.product-main .product-description')[0].text.strip() if soup.select('.product-main .product-description') != [] else ''
        desc_data = soup.select('.product-main .product-description')
        if desc_data != []:
            sheet1.cell(data_row,9).value = str(desc_data)
        Striked_Price = soup.select('.product-main .price--compare-at span.money')[0].text.strip().replace('₱', '') if soup.select('.product-main .price--compare-at span.money') != [] else ''
        prod_img = soup.select('.product-gallery--image-background img')[0]['src'] if soup.select('.product-gallery--image-background img') != [] else ''
        product_data = soup.select('script.analytics')
        if product_data != []:
            prod_dict = re.findall(r'window\.ShopifyAnalytics\.lib\.track\("Viewed Product",(.*?)\);',str(product_data)) if len(re.findall(r'window\.ShopifyAnalytics\.lib\.track\("Viewed Product",(.*?)\);',str(product_data))) != 0 else ''
            if len(prod_dict) != 0:
                dict_data = prod_dict[0]
                sku_data = re.findall(r'"sku":"(.*?)",',str(dict_data)) if re.findall(r'"sku":"(.*?)",',str(dict_data)) != [] else ''
                if len(sku_data) != 0:
                    prod_sku = sku_data[0]
                    sheet1.cell(data_row,6).value = prod_sku
                brand_data = re.findall(r'"brand":"(.*?)",',str(dict_data)) if re.findall(r'"brand":"(.*?)",',str(dict_data)) != [] else ''
                if len(brand_data) != 0:
                    prod_brand = brand_data[0].strip()
                    sheet1.cell(data_row,5).value = prod_brand
        print(row_num-1,'----->',input_url)
        sheet1.cell(data_row,1).value = input_url
        sheet1.cell(data_row,2).value = prod_title
        sheet1.cell(data_row,3).value = prod_price
        sheet1.cell(data_row,4).value = Striked_Price
        sheet1.cell(data_row,7).value = prod_img
        sheet1.cell(data_row,8).value = prod_desc
        data_row += 1
        try:
            wb.save('Output_Handyman.xlsx')
        except:
            wb.save('Output_Handyman.xlsx')
end_time = datetime.now()
total_time = end_time - start_time
print('----------------------------------------------------------------------------------------------------')
print('Script Start Time : ',start_time)
print('Script End Time : ',end_time)
print('Total Time : ', total_time)
try:
    wb.save('Output_Handyman.xlsx')
except:
    wb.save('Output_Handyman.xlsx')
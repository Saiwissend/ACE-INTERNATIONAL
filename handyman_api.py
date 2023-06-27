import requests
import json
import openpyxl
from datetime import datetime
from os import getcwd,path
import urllib.parse
wb = openpyxl.Workbook()
data_sheet = wb.active
data_sheet["A1"] = "product_url"
data_sheet["B1"] = "id_of_product"
data_sheet["C1"] = "product_name"
data_sheet["D1"] = "old_price (₱) "
data_sheet["E1"] = "sale_price (₱)"
data_sheet["F1"] = "image_url"

file_path = getcwd() 
file_name = 'Handyman'
max_row = 2
url = "https://pre-prod-api.gocart.ph/s5/api/v1/web/product/list"
for num in  range(1,64):
    payload = json.dumps({
    "buId": "3",
    "pageNo":f'{num}',
    "pageSize": "100",
    "sortBy": "0",
    "storeCode": "101",
    "zoneNum": "14",
    })
    headers = {
    'Accept': '*/*',
    'Accept-Language': 'en-US,en;q=0.9',
    'Connection': 'keep-alive',
    'Content-Type': 'application/json',
    'Origin': 'https://gocart.ph',
    'Referer': 'https://gocart.ph/',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-site',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
    'client': 'WEB',
    'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"'
    }

    response = requests.request("POST", url, headers=headers, data=payload,verify=False)

    json_validate = json.loads(response.text)
    if json_validate['data'] !=[]:
        for collect_id in json_validate['data']:
            id_of_product = str(collect_id['id'])
            product_name= collect_id['name']
            stock_status = collect_id['stockStatus']
            image_url = collect_id['image']
            old_price = collect_id['originalPrice']
            sale_price = (collect_id['price'])
            title_lower = product_name.lower().replace(' ','-').replace('#','21')

            product_url = f'https://gocart.ph/handyman/product/{title_lower}?buId=3&productId={id_of_product}&storeCode=101&zoneNum=14&isGoSubscribe=false'
            # product_payload = json.dumps({
            #     "buId": "3",
            #     "productId": id_of_product,
            #     "storeCode": "101",
            #     "zoneNum": "14",
            #     "isGoSubscribe": "false",
            #     "storeName": "handyman-do-it-best",
            #     "productName": product_name.lower().replace(' ','-')
            #     })
            # response_product = requests.request("POST", product_url, headers=headers, data=product_payload,verify=False)
            data_sheet.cell(row =max_row, column =1).value =product_url
            data_sheet.cell(row =max_row, column =2).value =id_of_product
            data_sheet.cell(row =max_row, column =3).value =product_name
            data_sheet.cell(row =max_row, column =4).value =old_price
            data_sheet.cell(row =max_row, column =5).value =sale_price
            data_sheet.cell(row =max_row, column =6).value =image_url
            max_row+=1
            
    try:
        wb.save(f"{file_path}\\{file_name}_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
    except:
        print(('Please Close the file'))
        wb.save(f"{file_path}\\{file_name}_{datetime.now().strftime('%d_%m_%Y')}.xlsx")

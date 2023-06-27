import re
import scrapy
import warnings
import openpyxl
from os import getcwd,path
from datetime import datetime
from bs4  import BeautifulSoup
import os, os.path, time, json, pathlib
from datetime import datetime
import sqlite3 as sql
import pandas as pd
from deep_translator import GoogleTranslator


class CemacooSpider(scrapy.Spider):
    name = "cemaco"
    start_urls = ['https://www.cemaco.com/sitemap.xml']
    now_time = time.localtime()
    now_date = time.strftime("%d-%m-%Y", now_time)
    time_on = time.strftime("%I:%M %p", now_time)
    timer_start = time.time()
    current_date = datetime.now().strftime('%d_%m_%Y')
    current_timestamp = datetime.now()

#---------------------------------------------------- 
    wb = openpyxl.Workbook()
    data_sheet = wb.active
    data_sheet.title = 'Product_collection'
    data_sheet["A1"] = "Product Url"
    data_sheet["B1"] = "Title"
    data_sheet["C1"] = "Sku"
    data_sheet["D1"] = "Brand"
    data_sheet["E1"] = "Price (Q)"
    data_sheet["F1"]= "Striked Price (Q)"
    data_sheet["I1"] = "Image"
  
    max_row = 2
    file_path = getcwd() 
    def clean(self,text):
        '''remove extra spaces & junk character'''
        text = re.sub(r'\n+','',text)
        text = re.sub(r'\s+',' ',text)
        text = re.sub(r'\r+',' ',text)
        return text.strip()

    def parse(self, response):
        if re.findall(r'<loc>(.*?)<\/loc>',response.text):
            url_collection = re.findall(r'<loc>(.*?)<\/loc>',response.text)
            for i in url_collection:
                if 'product' in i :
                    print('DOMAIN',i)
                    yield scrapy.Request(i,callback=self.parse_detail,dont_filter=True)
    def parse_detail(self,response):
        if re.findall(r'<loc>(.*?)<\/loc>',response.text):
            url_collection = re.findall(r'<loc>(.*?)<\/loc>',response.text)
            for url in url_collection:                
                yield response.follow(url,callback=self.listing_product)
                
    def listing_product(self,response):
        if response.xpath('//div[contains(@class,"productSkuName")]/text()'):
            item = {}
            item['product_url'] = response.url
            title = response.xpath('//div[contains(@class,"productSkuName")]/text()').get('').strip()
            if title:
                
                item['title'] = GoogleTranslator(source='auto', target='en').translate(text=title)
            else:
                item['title'] = ''
            
            item['sku'] = response.xpath('//meta[@property="product:sku"]/@content').get('').strip()
            item['image'] = response.xpath('//meta[@property="og:image"]/@content').get('').strip()
            # item['price'] = response.xpath('//meta[@property="product:price:amount"]/@content').get('').strip()
            item['price'] =  response.xpath('//div/@data-price').get('').strip()
            item['striked_price'] =  ''.join(response.xpath('//div[contains(@class,"pdpMainInfo")]//span[contains(text(),"Reg:")]//text()').getall()).replace('Reg','').replace(':','').replace('Q','').strip()
            item['brand'] = response.xpath('//meta[@property="product:brand"]/@content').get('').strip()
            
            yield item
        else:
            yield scrapy.Request(response.url,callback=self.listing_product)
        # self.data_sheet.cell(row =self.max_row, column =1).value =product_url
        # self.data_sheet.cell(row =self.max_row, column =2).value =title       
        # self.data_sheet.cell(row =self.max_row, column =3).value =sku
        # self.data_sheet.cell(row =self.max_row, column =4).value =brand
        # self.data_sheet.cell(row =self.max_row, column =5).value =price
        # self.data_sheet.cell(row =self.max_row, column =6).value =''
        # self.data_sheet.cell(row =self.max_row, column =7).value =image
        # self.max_row +=1
        # try:
        #     print('----------------------Saving Excel')
        #     self.wb.save(f"{self.file_path}\\Cemacoo_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
        # except:
        #     print(('Please Close the file'))
        #     self.wb.save(f"{self.file_path}\\Cemacoo_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 